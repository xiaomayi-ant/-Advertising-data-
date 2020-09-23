import pandas as pd
import numpy as  np
import matplotlib.pyplot as plt
import seaborn as sns

#数据提取
data=pd.read_excel(r'C:\Users\ant.zheng\Desktop\test.xlsx',index_col='Date')

#描述
print(data.head(10))
pd.set_option('display.max_columns',None)
print(data.describe())       #存在缺失值，数据分布右偏

#数据清洗
print(data.info())
print(data.isnull().sum())
print(data[data.isnull().values==True])  #查询缺失值行列，此条数据选择删除
data=data.dropna(axis=0)
data=data.drop_duplicates(keep='first')

#数据分布
#创建直方图
fig,ax=plt.subplots(1,1,figsize=(8,5))
ax.hist(data['Install'],bins=20)
plt.show()
#创建箱型图
columns=data.columns.values.tolist()
col=columns[3:7]
figure=plt.figure(figsize=(10,8))
for i,j in enumerate(col):#获取列字段
        ax=figure.add_subplot(2,2,i+1)
        j=str(j)
        data.boxplot(column=j,ax=ax)
plt.show()
#相关性描述(各项指标高度相关)
cor=data[['Install','Click','Spend','Impression']]
corre=cor.corr()                            #默认pearson
cmap='Set1'
sns.heatmap(corre,cmap=plt.get_cmap(cmap))
plt.show()

#生命周期曲线
data['Install'].plot()
data['Spend'].plot()
plt.show()

#考虑到本组数据之间相关性极高，用vif进行分析
#方差膨胀因子
from statsmodels.formula.api  import ols
def vif(data,col_i):
    """
    data:整份数据
    col_i:被检测的列名
    """
    cols=list(data.columns)
    cols.remove(col_i)
    cols_noti=cols
    formula=col_i+'~'+'+'.join(cols_noti)
    r2=ols(formula,data).fit().rsquared
    return 1./(1.-r2)
test_data=data[['Click','Impression','Spend']]
for i in test_data.columns:
    print(i,'\t',vif(test_data,i))

"""我们希望看到这些维度数据的信息，因此未选择删除其中一些字段，尝试Ridge"""
from sklearn.model_selection  import train_test_split
from sklearn import linear_model
from scipy import stats
import sklearn.preprocessing  as preprocessing
import openpyxl

#正态性检测
u=data['Install'].mean()
std=data['Install'].std()
result=stats.kstest(data['Install'],'norm',(u,std))
print("正太检验结果：")
print(result)

#数据归一化
def Normalization():
    #对数据进行归一化处理，并存储到
    data=pd.read_excel(r'C:\Users\ant.zheng\Desktop\test.xlsx')
    sam=[]
    a=['Click','Spend','Impression']
    for  i  in a :
        y=data.loc[:,i]
        ys=list(preprocessing.scale(y))  #归一化
        sam.append(ys)
    sam.append(data['Install'])
    return sam

def write_excel_xlsx(path,sheet_name,value):
    index=len(value)
    workbook=openpyxl.Workbook()
    sheet=workbook.active
    sheet.title=sheet_name
    field=['Click','Spend','Impression','Install']
    for f in range(0,len(field)):
        sheet.cell(row=1,column=f+1,value=str(field[f]))
    for i  in range(1,index):
        for j in range(0,len(value[i-1])):
            sheet.cell(row=i+1,column=j+1,value=str(value[i-1][j]))
    workbook.save(path)
    print('xlsx导入数据成功！')

#Ridge回归
def load_data():
        raw=pd.read_excel(r'C:\Users\ant.zheng\Desktop\dataset\Regression\norm.xlsx')
        raw=raw.dropna()
        raw=raw.drop_duplicates()
        raw.data=raw[['Click','Spend','Impression']]
        raw.target=raw['Install']
        return train_test_split(raw.data,raw.target,test_size=0.3,random_state=1)

def test_ridge(*data):
        X_train,X_test,y_train,y_test=data
        ridgeRegression=linear_model.Ridge()
        ridgeRegression.fit(X_train,y_train)
        print("权重向量:%s, b的值为:%.2f" % (ridgeRegression.coef_, ridgeRegression.intercept_))
        print("损失函数的值:%.2f" % np.mean((ridgeRegression.predict(X_test) - y_test) ** 2))
        print("预测性能得分: %.2f" % ridgeRegression.score(X_test, y_test))
        #查看test数据拟合效果
        plt.figure(figsize=(10, 8))
        res=pd.Series(ridgeRegression.predict(X_test).tolist())
        y=pd.Series(y_test.tolist())
        res.plot()
        y.plot()
        plt.show()

if __name__ == '__main__':
    value = Normalization()
    value = np.array(value).T
    path = r'C:\Users\ant.zheng\Desktop\dataset\Regression\norm.xlsx'
    sheet_name = '归一化'
    write_excel_xlsx(path, sheet_name, value)
    X_train, X_test, y_train, y_test = load_data()
    test_ridge(X_train, X_test, y_train, y_test)








